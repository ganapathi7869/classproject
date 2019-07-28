import click
from openpyxl import load_workbook
import os

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'classproject.settings')

import django
django.setup()

from onlineapp.models import *

''''@click.group()
def main():
    pass

@main.command()'''
def importdatacolleges():
    wb=load_workbook('students.xlsx')
    ws=wb.get_sheet_by_name('Colleges')
    r=ws.max_row
    c=ws.max_column
    for i in range(2,r+1):
        col=[]
        for j in range(1,c+1):
            col.append(ws.cell(row=i,column=j).value)
        inst=College(name=col[0],location=col[2],acronym=col[1],contact=col[3])
        inst.save()

def importdatastudents():
    wb=load_workbook('students.xlsx')
    ws=wb.get_sheet_by_name('Current')
    r = ws.max_row
    c = ws.max_column
    for i in range(2, r + 1):
        stu = []
        for j in range(1, c + 1):
            stu.append(ws.cell(row=i, column=j).value)
        inst = Student(name=stu[0], email=stu[2],db_folder=stu[3],college=College.objects.get(acronym=stu[1]))
        inst.save()

def importdatadropoutstudents():
    wb=load_workbook('students.xlsx')
    ws=wb.get_sheet_by_name('Deletions')
    r = ws.max_row
    c = ws.max_column
    for i in range(2, r + 1):
        stu = []
        for j in range(1, c + 1):
            stu.append(ws.cell(row=i, column=j).value)
        inst = Student(name=stu[0],dropped_out=True, email=stu[2],db_folder=stu[3],college=College.objects.get(acronym=stu[1]))
        inst.save()

def importdatamocktest1():
    wb = load_workbook('outputexcel.xlsx')
    ws = wb.get_sheet_by_name('Sheet')
    r = ws.max_row
    c = ws.max_column
    for i in range(2, r + 1):
        stu = []
        for j in range(1, c + 1):
            if j==1:
                # print(type(ws.cell(row=i, column=j).value))
                string=ws.cell(row=i, column=j).value.split('_')[2]
                stu.append(string)
            else:
                stu.append(ws.cell(row=i, column=j).value)
        inst = MockTest1(student=Student.objects.get(db_folder=stu[0]),problem1=stu[1],problem2=stu[2],problem3=stu[3],problem4=stu[4],total=stu[5])
        inst.save()

def getcolleges():
    c=College.objects.all()
    for ob in c:
        print(ob.name,ob.location,ob.acronym,ob.contact)

def getcollegescount():
    return College.objects.count()

def getcollegesacronymcontact():
    # for i in College.objects.all():
    #     print(i.acronym,i.contact)

    for i in College.objects.values('acronym','contact'):
        print(i)
# print(  getcollegesacronymcontact()  )

def getcollegescountlocationwise():
    loc='Hyderabad'
    count=0
    # print(type(College.objects.values('location')))
    for i in College.objects.values('location'):
        if i['location'] == 'Hyderabad':
            count+=1
    print(count)

    # print( College.objects.filter(location='Hyderabad').count())  #this is optimal
# print(getcollegescountlocationwise())

def getsortedcollegesacronymwise():
    for i in College.objects.order_by('acronym'):
        print(i)

def getsortedcollegesacronymwisedesc():
    for i in College.objects.order_by('-acronym'):
        print(i)

def getcollegesdesctop5():
    for i in College.objects.order_by('-acronym')[:5]:
        print(i)

def getcollegesdescbottom5():
    # print(type(College.objects.order_by('acronym')))
    for i in College.objects.order_by('acronym')[:5]:
        print(i)

from django.db.models import *


def getcollegestats():
    for i in College.objects.values('location').annotate(count=Count('location')):
        print(i)
# getcollegestats()

def getcollegestatsordered():
    for i in College.objects.values('location').annotate(count=Count('location')).order_by('count'):
        print(i)

def getcollegewisedroppedoutstusents():
    for i in Student.objects.filter(dropped_out=True).values('college').distinct():
        print (i)

def getcollegewisenoofstudents():
    for i in Student.objects.values('college__acronym').annotate(count=Count('id')):
        print(i)

def getlocationwisenoofstudents():
    for i in Student.objects.values('college__location').annotate(count=Count('id')).order_by('-count'):
        print(i)

def getlocationofmaxstudents():
    print(Student.objects.values('college__location').annotate(count=Count('id')).order_by('-count'))
print(getlocationofmaxstudents())

def getstudentcollegemarks():
    # for i in MockTest1.objects.values('student__name','total','student__college__acronym'):
    #     print(i)

    #met 2
    for i in Student.objects.values('name','mocktest1__total','college__acronym'):
        print (i)

def getstudentsmarksgte30():
    for i in MockTest1.objects.values('student__name','total','student__college__acronym').filter(total__gte=30):
        print(i)

def getstudentcountmarksgte30():
    return(MockTest1.objects.values('total').filter(total__gte=30).count())
# print(  getstudentcountmarksgte30()  )
